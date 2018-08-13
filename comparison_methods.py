import time
import math
import ReadData as RD
import Transform as Tf
import Locate as Lc
import StructMesh as SM
import GetZ as GZ
import GetThick as GT
from openpyxl import Workbook
import numpy as np

start = time.time()
Point = RD.vertices('data/input/sh_0411.obj')
P = Lc.locate(Point)
face1 = []
face2 = []
face1_r = []
face2_r = []
for point in P['LW']:
    face1.append(point)
    face1_r.append(point)
for point in P['RW']:
    face2.append(point)
    face2_r.append(point)

face1.sort()
face2.sort()

m = 4
n = 1
corner = [[-98.7, -178.03], [61.3, -178.03], [-98.7, -168.05], [61.3, -168.05]]

#m = 9
#n = 9
#corner = [[-150, -140], [136, -140], [-150, 150], [136, 150]]

t = SM.rect(corner, m, n)

GZ.getz(t, face1, face2, 2, 20)
GT.thick(t, face1, face2, 3, 20)

#rotate model
ang = math.pi / 9
Tf.rotate_x(face1_r, ang)
Tf.rotate_x(face2_r, ang)

face1_r.sort()
face2_r.sort()

points = [[t['vertices'][i][0], t['vertices'][i][1], t['z'][i]] for i in range(len(t['vertices']))]
Tf.rotate_x(points, ang)
t_r = {'vertices':list([point[0], point[1]] for point in points)}
z0 = [point[2] for point in points]

GZ.getz(t_r, face1_r, face2_r, 2, 20)
GT.thick(t_r, face1_r, face2_r, 3, 20)

#write data to excel
wb = Workbook()
ws1 = wb.active
ws1.title = 'vertices'
ws1.append(['corner', 'x', 'y'])
for i in range(len(corner)):
    ws1.append([i+1, corner[i][0], corner[i][1]])
ws1.append([])

ws1.append(['m', 'n'])
ws1.append([m, n])
ws1.append([])

ws1.append(['point', 'x', 'y', 'z', 'z0', 'error', 'R1', 'a1', 'b1', 'R2', 'a2', 'b2'])
for i in range(len(t['vertices'])):
    ws1.append([i+1, t['vertices'][i][0], t['vertices'][i][1], t['z'][i], z0[i], abs(100*(t_r['z'][i]-z0[i]) / t['thick2'][i]), t['R1'][i],
    t['a1'][i], t['b1'][i], t['R2'][i], t['a2'][i], t['b2'][i]])
    ws1.append([i+1, t_r['vertices'][i][0], t_r['vertices'][i][1], t_r['z'][i], z0[i], abs(100*(t_r['z'][i]-z0[i]) / t['thick2'][i]),
    t_r['R1'][i], t_r['a1'][i], t_r['b1'][i], t_r['R2'][i], t_r['a2'][i], t_r['b2'][i]])
ws1.append([])

ws2 = wb.create_sheet(title = 'thick')
ws2.append(['point'])
for row in range(1,2):
    for col in range(2, len(t['vertices']) + 2):
        ws2.cell(row = row, column = col, value = col-1)

T1 = ['thick1']
T1.extend(t['thick1'])
ws2.append(T1)

T1_r = ['thick1_r']
T1_r.extend(t_r['thick1'])
ws2.append(T1_r)

T1_error = ['thick1_error/%']
T1_error.extend(100 * abs(np.array(t_r['thick1']) - np.array(t['thick1'])) / np.array(t['thick1']))
ws2.append(T1_error)
ws2.append([])

T2 = ['thick2']
T2.extend(t['thick2'])
ws2.append(T2)

T2_r = ['thick2_r']
T2_r.extend(t_r['thick2'])
ws2.append(T2_r)

T2_error = ['thick2_error/%']
T2_error.extend(100 * abs(t_r['thick2'] - t['thick2']) / t['thick2'])
ws2.append(T2_error)
ws2.append([])

wb.save('data/output/thickness_comparison.xlsx')
end = time.time()
print('Time=',end - start)
